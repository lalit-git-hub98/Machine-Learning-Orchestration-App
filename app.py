import streamlit as st
from streamlit_modal import Modal
import numpy as np
import pandas as pd
import mlflow
from PIL import Image
import sklearn
import os
import shutil
from zipfile import ZipFile
import ast
import pickle
import joblib

from sklearn.model_selection import train_test_split

from openpyxl import Workbook
from openpyxl import load_workbook

import tensorflow as tf
from tensorflow.keras.utils import to_categorical
from tensorflow.keras.models import Sequential, load_model
from tensorflow.keras.layers import Conv2D, MaxPool2D, Dense, Flatten, Dropout

from sklearn.preprocessing import LabelEncoder
from sklearn.preprocessing import StandardScaler

from sklearn.decomposition import PCA

from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import PolynomialFeatures
from sklearn.linear_model import Lasso

from sklearn.linear_model import LogisticRegression
from sklearn.tree import DecisionTreeClassifier
from sklearn.ensemble import RandomForestClassifier
from xgboost import XGBClassifier
from sklearn.naive_bayes import GaussianNB

from sklearn.metrics import mean_absolute_error
from sklearn.metrics import mean_squared_error
from sklearn.metrics import r2_score

from sklearn.metrics import accuracy_score
from sklearn.metrics import precision_score
from sklearn.metrics import recall_score
from sklearn.metrics import f1_score
from sklearn.metrics import confusion_matrix



st.set_page_config(page_title = 'ML Orchestration App')

st.markdown("<h1 style='text-align:center; color:white;'>Machine Learning Orechestration App</h1>", unsafe_allow_html = True)

if 'file_check' not in st.session_state:
    st.session_state.file_check = False

if st.session_state.file_check == False:
    ################## CSV files check #############################
    if len(os.listdir('./models_csv')) != 0:
        path1 = './models_csv'
        for i in os.listdir(path1):
            os.remove(os.path.join(path1, i))
    if len(os.listdir('./Parameters_Track_CSV')) != 0:
        path1 = './Parameters_Track_CSV'
        for i in os.listdir(path1):
            os.remove(os.path.join(path1, i))
    if len(os.listdir('./Experiment_Track_CSV')) != 0:
        path1 = './Experiment_Track_CSV'
        for i in os.listdir(path1):
            os.remove(os.path.join(path1, i))
    #################### Image files check ##############################
    if len(os.listdir('./models_image')) != 0:
        path1 = './models_image'
        for i in os.listdir(path1):
            os.remove(os.path.join(path1, i))
    if len(os.listdir('./Parameters_Track_Image')) != 0:
        path1 = './Parameters_Track_Image'
        for i in os.listdir(path1):
            os.remove(os.path.join(path1, i))
    if len(os.listdir('./Experiment_Track')) != 0:
        path1 = './Experiment_Track'
        for i in os.listdir(path1):
            os.remove(os.path.join(path1, i))
    st.session_state.file_check = True

current_dir = os.getcwd()
data_directory = 'imageData'

data_list = ['CSV', 'Image']
data_select = st.selectbox('Select Data Type', data_list)

if 'image_data1' not in st.session_state:
    st.session_state.image_data1 = None
if 'image_labels1' not in st.session_state:
    st.session_state.image_labels1 = None
if 'num_of_classes1' not in st.session_state:
    st.session_state.num_of_classes1 = None


if 'display_tensor_dash' not in st.session_state:
    st.session_state.display_tensor_dash = None

st.header('Folder/File Upload')

if data_select == 'Image':
    st.session_state.display_model_dash = False
    file_upload_options = ['Upload Folder', 'Use Default Folder']
    upload_options = st.selectbox('Select File Upload Option', file_upload_options)
    if upload_options == 'Upload Folder':
        uploaded_files = st.file_uploader("Upload a folder", type="zip", accept_multiple_files=False)
    else:
        upload_folder_name = './upload_files_store/Train.zip'
        #if st.button('Confirm Data', key = 'upload_button'):
        uploaded_files = upload_folder_name
        st.write('The default data is traffic sign classifier dataset. It has 43 classes.')


    def dataPreprocessing(data_directory, img_size=30):
        image_data = []
        image_labels = []
        e_n = 0
        num_of_classes = 0
        for i in os.listdir(data_directory):
            class_images = os.path.join(data_directory, i)
            st.write(f'Processing Images from class', {i})
            for image in os.listdir(class_images):
                try:
                    path_1 = os.path.join(class_images, image)
                    #print(path_1)
                    image = Image.open(path_1).convert('L')
                    image = image.resize((img_size,img_size))
                    # Resizing all images into 30*30
                    image =np.array(image)
                    image_data.append(image)
                    image_labels.append(num_of_classes)
                except:
                    e_n += 1
            num_of_classes += 1
        st.write(e_n, ' Images were not processed from!!!')
        image_data = np.array(image_data)
        image_labels = np.array(image_labels)
        return image_data, image_labels, num_of_classes

    #@st.cache_resource(experimental_allow_widgets=True)
    def up_files(): 
        data_path = os.path.join(current_dir, data_directory)

        if os.path.exists(data_path):
            shutil.rmtree(data_path)
        os.mkdir(data_path)

        if os.path.exists('./models_csv'):
            shutil.rmtree('./models_csv')
        os.mkdir('./models_csv')

        if os.path.exists('./models_image'):
            shutil.rmtree('./models_image')
        os.mkdir('./models_image')


        if uploaded_files is not None:
            try:
                with open(uploaded_files.name, "wb") as f:
                    f.write(uploaded_files.read())
                for i in os.listdir(current_dir):
                    if i.endswith('.zip'):
                        zip_file = str(i)
                zip_path = os.path.join(current_dir, zip_file)
                with ZipFile(zip_path, 'r') as zObject:
                    zObject.extractall(path=data_path)
            except:
                for i in os.listdir('./upload_files_store/'):
                    if i.endswith('.zip'):
                        #st.write(i)
                        zip_file = str(i)
                zip_path = os.path.join('./upload_files_store', zip_file)
                #st.write(zip_path)
                with ZipFile(zip_path, 'r') as zObject:
                    zObject.extractall(path=data_path)
            
            #os.system(f"unzip {uploaded_files.name} -d {data_path}")
                
            for k in os.listdir('imageData'):
                archive_data_dir = k
                break
            
            data_path1 = os.path.join(current_dir, data_directory)
            data_path2 = os.path.join(data_path1, archive_data_dir)

            for i in os.listdir(data_path2):
                path_1 = os.path.join(data_path2, i)
                for j in os.listdir(path_1):
                    path_2 = os.path.join(path_1, j)
                    image = Image.open(path_2)
                    image = image.resize((30,30))
                    image =np.array(image)
                    print(image.shape)
                    break
            st.session_state.image_data1, st.session_state.image_labels1, st.session_state.num_of_classes1 = dataPreprocessing(data_path2)
            st.success("Folder uploaded successfully!")

            for i in os.listdir(current_dir):
                if i.endswith('.zip'):
                    zip_file = i
                    os.remove(os.path.join(current_dir, zip_file))

    col_btn1, col_btn2 = st.columns(2)

    with col_btn1:
        if st.button('Confirm Data'):
            st.session_state.layers_track_dict = {}
            up_files()
            st.session_state.display_tensor_dash = True

    with col_btn2:
        if st.button('Remove File'):
            data_path = os.path.join(current_dir, data_directory)

            if os.path.exists(data_path):
                shutil.rmtree(data_path)
            os.mkdir(data_path)

            st.session_state.display_tensor_dash = False

            st.success("Data Removed Successfully!")

    st.header('Validation Split')
    test_split_perc = st.number_input('Select validation split percentage(*100)', min_value = 0.1, max_value = 0.9, step = 0.01)

    if st.session_state.display_tensor_dash == True:
        st.session_state.display_model_dash = False
        layer_list = ['Conv2D', 'MaxPool2D', 'Dropout', 'Flatten', 'Dense']
        activations_list = ['relu', 'gelu', 'softmax', 'linear', 'sigmoid', 'tanh', 'softplus']

        if 'layers_track_dict' not in st.session_state:
            st.session_state.layers_track_dict = {}

        st.header('Layer Setup')

        layer_select = st.selectbox('Select Layer Type', layer_list)

        if layer_select == 'Conv2D':
            col1, col2, col3 = st.columns(3)
            ####### Setting Columns ###############
            with col1:
                no_of_filters = st.number_input('No of Filters', min_value = 2, step = 1)
            with col2:
                kernel_size = st.number_input('Kernel Size', min_value = 2, step = 1)
            with col3:
                activation = st.selectbox('Activation Function', activations_list)
            ######### Storing Values #################
            layer_dict = {}
            layer_dict['layer'] = 'Conv2D'
            layer_dict['filters'] = no_of_filters
            layer_dict['kernelSize'] = kernel_size
            layer_dict['Activation'] = activation


        elif layer_select == 'MaxPool2D':
            ####### Setting Columns ###############
            pool_size = st.number_input('Pool Size', min_value = 2, step = 1)
            ######### Storing Values #################
            layer_dict = {}
            layer_dict['layer'] = 'MaxPool2D'
            layer_dict['poolSize'] = pool_size

        elif layer_select == 'Dropout':
            ####### Setting Columns ###############
            dropout_rate = st.number_input('Dropout Rate', min_value = 0.0, max_value = 1.0, step = 0.01)
            ######### Storing Values #################
            layer_dict = {}
            layer_dict['layer'] = 'Dropout'
            layer_dict['dropoutRate'] = '{0:.2f}'.format(dropout_rate)

        elif layer_select == 'Flatten':
            ####### Setting Columns ###############
            st.write('No parameters for Flatten layer.')
            ######### Storing Values #################
            layer_dict = {}
            layer_dict['layer'] = 'Flatten'

        elif layer_select == 'Dense':
            col1, col2 = st.columns(2)
            ####### Setting Columns ###############
            with col1:
                units = st.number_input('No of Dense Units', min_value = 1, step = 1)
            with col2:
                activation = st.selectbox('Activation Function', activations_list)
            ######### Storing Values #################
            layer_dict = {}
            layer_dict['layer'] = 'Dense'
            layer_dict['Units'] = units
            layer_dict['Activation'] = activation

        else:
            st.write('Please select a layer to proceed!!!')

        col1, col2, col3 = st.columns(3)
        with col1:
            if st.button('Add Layer'):
                key_no = len(st.session_state.layers_track_dict) + 1
                st.session_state.layers_track_dict[key_no] = layer_dict
        with col2:
            if st.button('Remove Last Layer'):
                try:
                    last_key = list(st.session_state.layers_track_dict)[-1]
                    del st.session_state.layers_track_dict[last_key]
                except:
                    pass
        with col3:
            if st.button('Remove All Layers'):
                try:
                    st.session_state.layers_track_dict = {}
                except:
                    pass

        # if len(st.session_state.layers_track_dict) > 1:
        #     cols1, cols2 = st.columns(2)
        #     col_to_remove = 1
        #     with cols1:
        #         col_to_remove = st.number_input('Layer No', min_value = 1, max_value = len(st.session_state.layers_track_dict), step = 1)
        #     with cols2:
        #         st.write(' ')
        #         st.write(' ')
        #         if st.button('Remove Layer'):
        #                 try:
        #                     remove_key = list(st.session_state.layers_track_dict)[col_to_remove - 1]
        #                     del st.session_state.layers_track_dict[remove_key]
        #                 except:
        #                     pass


        #st.write(layer_select)
                ############################################
        st.write(st.session_state.layers_track_dict)

        
        # st.write(X_train.shape, X_test.shape, y_train.shape, y_test.shape)

        # st.write(len(st.session_state.layers_track_dict))

        ###############################################################

        st.header('Compile Model')
    
        optimizers_list = ['Adadelta', 'Adagrad', 'adam', 'AdamW', 'RMSprop', 'SGD']
        loss_fn_list = ['categorical_crossentropy', 'CategoricalFocalCrossentropy', 'CosineSimilarity', 'SparseCategoricalCrossentropy', 
                        'Poisson', 'Mean Absolute Error', 'Mean Squared Error', 'Mean Squared Logarithmic Error', 'Huber']
        metrics_list = ['Accuracy', 'Precision', 'Recall', 'True Positives', 'False Positives', 'True Negatives', 'False Negatives']

        col_compile_1, col_compile_2 = st.columns(2)

        with col_compile_1:
            select_optimizer = st.selectbox('Select Optimizer', optimizers_list)
        with col_compile_2:
            select_loss = st.selectbox('Select Loss Function', loss_fn_list)
        select_metrics = st.multiselect('Select Metrics', metrics_list)

        select_metrics_keras = []
        for i in range(len(select_metrics)):
            if select_metrics[i] == 'Accuracy':
                select_metrics_keras.append(tf.keras.metrics.Accuracy())
            elif select_metrics[i] == 'Precision':
                select_metrics_keras.append(tf.keras.metrics.Precision())
            elif select_metrics[i] == 'Recall':
                select_metrics_keras.append(tf.keras.metrics.Recall())
            elif select_metrics[i] == 'True Positives':
                select_metrics_keras.append(tf.keras.metrics.TruePositives())
            elif select_metrics[i] == 'False Positives':
                select_metrics_keras.append(tf.keras.metrics.FalsePositives())
            elif select_metrics[i] == 'True Negatives':
                select_metrics_keras.append(tf.keras.metrics.TrueNegatives())
            elif select_metrics[i] == 'False Negatives':
                select_metrics_keras.append(tf.keras.metrics.FalseNegatives())
        

        st.header('Training Parameters')
        cols_train_params1, cols_train_params2 = st.columns(2)

        with cols_train_params1:
            no_epochs = st.number_input('Epochs', min_value = 1, step = 1)
        with cols_train_params2:
            batch_size = st.number_input('Batch Size', min_value = 1, step = 1)

        st.header('Experiment Name')
        if 'model_name' not in st.session_state:
            st.session_state.model_name = None
        st.session_state.model_name = st.text_input('Enter the experiment name')

        # csv_metrics = ['Loss', 'Val_Loss']

        if len(os.listdir('./Experiment_Track')) == 0:
            workbook = Workbook()
            worksheet = workbook.active

            col_names_list = ['No', 'Model Name', 'Loss', 'Val_Loss', 'Accuracy', 'Val_Accuracy', 
                                'Precision', 'Val_Precision', 'Recall', 'Val_Recall', 'True Positives', 'Val_True_Positives',
                                'True Negatives', 'Val_True_Negatives', 'False Positives', 'Val_False_Positives',
                                'False Negatives', 'Val_False_Negatives']
            
            row_number = 1

            for col, value in enumerate(col_names_list, start=1):
                worksheet.cell(row=row_number, column=col, value=value)

            workbook.save(filename='./Experiment_Track/Experiment_Tracker.xlsx')

        if len(os.listdir('./Parameters_Track_Image')) == 0:
            workbook = Workbook()
            worksheet = workbook.active

            col_names_list = ['No', 'Model Name', 'Validation Split', 'Architecture', 
                              'Optimizer', 'Loss Function', 'Epochs', 'Batch Size']
            
            row_number = 1

            for col, value in enumerate(col_names_list, start=1):
                worksheet.cell(row=row_number, column=col, value=value)

            workbook.save(filename='./Parameters_Track_Image/Parameters_Tracker_Image.xlsx')



        if st.button('Run Experiemnt'):
            X_train, X_test, y_train, y_test =train_test_split(st.session_state.image_data1, st.session_state.image_labels1, test_size=float(test_split_perc))
            X_train = X_train.reshape(X_train.shape[0], 30, 30, 1)
            X_test = X_test.reshape(X_test.shape[0], 30, 30, 1)

            y_train = to_categorical(y_train,st.session_state.num_of_classes1)
            y_test = to_categorical(y_test,st.session_state.num_of_classes1)

            st.write(y_train.shape)
            st.write(y_test.shape)

            ########################################################################################################################

            if 'model' not in st.session_state:
                st.session_state.model = Sequential()
            #model = Sequential()

            keysList = list(st.session_state.layers_track_dict.keys())
            #st.write(keysList)
            for i in range(len(st.session_state.layers_track_dict)):
                if i == 0:
                    st.session_state.model = Sequential()
                    if st.session_state.layers_track_dict[keysList[i]]['layer'] == 'Conv2D':
                        st.session_state.model.add(Conv2D(filters=int(st.session_state.layers_track_dict[keysList[i]]['filters']), 
                                         kernel_size=(int(st.session_state.layers_track_dict[keysList[i]]['kernelSize']), int(st.session_state.layers_track_dict[keysList[i]]['kernelSize'])), 
                                         activation=st.session_state.layers_track_dict[keysList[i]]['Activation'], 
                                         input_shape=X_train.shape[1:]))
                    else:
                        st.write('The first layer has to be Conv2D')
                        break
                else:
                    if st.session_state.layers_track_dict[keysList[i]]['layer'] == 'Conv2D':
                        st.session_state.model.add(Conv2D(filters=int(st.session_state.layers_track_dict[keysList[i]]['filters']), 
                                         kernel_size=(int(st.session_state.layers_track_dict[keysList[i]]['kernelSize']), int(st.session_state.layers_track_dict[keysList[i]]['kernelSize'])), 
                                         activation=st.session_state.layers_track_dict[keysList[i]]['Activation']))
                    elif st.session_state.layers_track_dict[keysList[i]]['layer'] == 'MaxPool2D':
                        st.session_state.model.add(MaxPool2D(pool_size=(int(st.session_state.layers_track_dict[keysList[i]]['poolSize']),int(st.session_state.layers_track_dict[keysList[i]]['poolSize']))))
                    elif st.session_state.layers_track_dict[keysList[i]]['layer'] == 'Dropout':
                        st.session_state.model.add(Dropout(rate=float(st.session_state.layers_track_dict[keysList[i]]['dropoutRate'])))
                    elif st.session_state.layers_track_dict[keysList[i]]['layer'] == 'Flatten':
                        st.session_state.model.add(Flatten())
                    elif st.session_state.layers_track_dict[keysList[i]]['layer'] == 'Dense':
                        st.session_state.model.add(Dense(int(st.session_state.layers_track_dict[keysList[i]]['Units']), 
                                        activation=(st.session_state.layers_track_dict[keysList[i]]['Activation'])))

            st.session_state.model.compile(loss=select_loss, optimizer=select_optimizer, metrics=select_metrics_keras)

            filename = './models_image/' + str(st.session_state.model_name) + '.h5'
            # with open(filename, 'wb') as model_file:
            #     joblib.dump(st.session_state.model, model_file)
            st.session_state.model.save(filename)



            #########################################################################################################################
            if st.session_state.model_name:
                workbook = load_workbook('./Experiment_Track/Experiment_Tracker.xlsx')
                worksheet = workbook.active

                non_empty_rows = 0

                for row in worksheet.iter_rows():
                    if any(cell.value is not None for cell in row):
                        non_empty_rows += 1

                col_names_list = []

                
                history = st.session_state.model.fit(X_train, y_train, batch_size=int(batch_size), epochs=int(no_epochs), validation_data=(X_test, y_test))
                st.write(history.history)
                col_names_list.append(non_empty_rows)
                col_names_list.append(str(st.session_state.model_name))
                col_names_list.append(history.history['loss'][-1])
                col_names_list.append(history.history['val_loss'][-1])

                if 'Accuracy' in select_metrics:
                    col_names_list.append(history.history['accuracy'][-1]) 
                    col_names_list.append(history.history['val_accuracy'][-1]) 
                else:
                    col_names_list.append('Nan') 
                    col_names_list.append('Nan')

                if 'Precision' in select_metrics:
                    col_names_list.append(history.history['precision'][-1]) 
                    col_names_list.append(history.history['val_precision'][-1]) 
                else:
                    col_names_list.append('Nan') 
                    col_names_list.append('Nan') 

                if 'Recall' in select_metrics:
                    col_names_list.append(history.history['recall'][-1]) 
                    col_names_list.append(history.history['val_recall'][-1]) 
                else:
                    col_names_list.append('Nan') 
                    col_names_list.append('Nan') 

                if 'True Positives' in select_metrics:
                    col_names_list.append(history.history['true_positives'][-1]) 
                    col_names_list.append(history.history['val_true_positives'][-1]) 
                else:
                    col_names_list.append('Nan') 
                    col_names_list.append('Nan') 

                if 'True Negatives' in select_metrics:
                    col_names_list.append(history.history['true_negatives'][-1]) 
                    col_names_list.append(history.history['val_true_negatives'][-1]) 
                else:
                    col_names_list.append('Nan') 
                    col_names_list.append('Nan') 

                if 'False Positives' in select_metrics:
                    col_names_list.append(history.history['false_positives'][-1]) 
                    col_names_list.append(history.history['val_false_positives'][-1]) 
                else:
                    col_names_list.append('Nan') 
                    col_names_list.append('Nan') 


                if 'False Negatives' in select_metrics:
                    col_names_list.append(history.history['false_negatives'][-1]) 
                    col_names_list.append(history.history['val_false_negatives'][-1]) 
                else:
                    col_names_list.append('Nan') 
                    col_names_list.append('Nan')  


                for col, value in enumerate(col_names_list, start=1):
                    worksheet.cell(row=non_empty_rows+1, column=col, value=value)
                
                workbook.save(filename='./Experiment_Track/Experiment_Tracker.xlsx')

                ############################# Saving Parameters #########################################
                workbook = load_workbook('./Parameters_Track_Image/Parameters_Tracker_Image.xlsx')
                worksheet = workbook.active

                non_empty_rows = 0

                for row in worksheet.iter_rows():
                    if any(cell.value is not None for cell in row):
                        non_empty_rows += 1

                col_params_list = []

                col_params_list.append(non_empty_rows)
                col_params_list.append(str(st.session_state.model_name))
                col_params_list.append(test_split_perc)
                col_params_list.append(str(st.session_state.layers_track_dict))
                col_params_list.append(select_optimizer)
                col_params_list.append(select_loss)
                col_params_list.append(no_epochs)
                col_params_list.append(batch_size)


                for col, value in enumerate(col_params_list, start=1):
                    worksheet.cell(row=non_empty_rows+1, column=col, value=value)
                
                workbook.save(filename='./Parameters_Track_Image/Parameters_Tracker_Image.xlsx')



                    
            else:
                st.write('Please provide name of the experiemnt to run the experiment')
        ###############################################################################################

        ############################ Create a Sidebar #####################################################
        try:
            st.markdown(
                """
                <style>
                    section[data-testid="stSidebar"] {
                        width: 30% !important; # Set the width to your desired value
                    }
                </style>
                """,
                unsafe_allow_html=True,
            )
            with st.sidebar:
                st.header('Experiments')
                if len(os.listdir('./Parameters_Track_Image')) == 0:
                    st.write('No experiments performed yet!!!')
                else:
                    wb_obj = load_workbook('./Parameters_Track_Image/Parameters_Tracker_Image.xlsx')
                    sheet_obj = wb_obj.active

                    total_rows = sheet_obj.max_row

                    pop_up_names = []
                    pop_up_val_splits = []
                    pop_up_architecture = []
                    pop_up_optimizer = []
                    pop_up_loss = []
                    pop_up_epochs = []
                    pop_up_batch = []

                    for u in range(2, total_rows+1):
                        pop_up_names.append(sheet_obj.cell(row = u, column = 2).value)
                        pop_up_val_splits.append(sheet_obj.cell(row = u, column = 3).value)
                        pop_up_architecture.append(sheet_obj.cell(row = u, column = 4).value)
                        pop_up_optimizer.append(sheet_obj.cell(row = u, column = 5).value)
                        pop_up_loss.append(sheet_obj.cell(row = u, column = 6).value)
                        pop_up_epochs.append(sheet_obj.cell(row = u, column = 7).value)
                        pop_up_batch.append(sheet_obj.cell(row = u, column = 8).value)

                
                
                    pop_col1, pop_col2, pop_col3, pop_col4 = st.sidebar.columns([0.1, 0.4, 0.25, 0.25])
                    with pop_col1:
                        st.write('No')
                    with pop_col2:
                        st.write('Model Name')
                    with pop_col3:
                        st.write('Check')
                    with pop_col4:
                        st.write('Download')

                    
                    for y in range(len(pop_up_names)):
                        modal = Modal(key=str(y),title=str(pop_up_names[y]))
                        for i in range(1):
                            with pop_col1:
                                st.write(str(y+1))
                                st.write(' ')
                            with pop_col2:
                                st.write(pop_up_names[y])
                                st.write(' ')
                            with pop_col3:
                                open_modal = st.button('Show', key=str(y))
                            with pop_col4:
                                y_str = y+1000
                                don_filename = './models_image/' + str(st.session_state.model_name) + '.h5'
                                display_filename = str(st.session_state.model_name) + '.h5'
                                with open(don_filename, 'rb') as f:
                                    st.download_button(':inbox_tray:', f, file_name=display_filename, key = str(y_str))  # :inbox_tray:  	:receipt:
                            
                        if open_modal:
                            with modal.container():
                                arc = ast.literal_eval(pop_up_architecture[y])
                                st.write('Name: ', pop_up_names[y])
                                st.write('Validation Split: ', pop_up_val_splits[y])
                                st.write('Optimizer: ', pop_up_optimizer[y])
                                st.write('Loss Function: ', pop_up_loss[y])
                                st.write('Epochs: ', pop_up_epochs[y])
                                st.write('Batch Size: ', pop_up_batch[y])
                                st.write('Architecture: ', arc)

            ml_results = pd.read_excel('./Experiment_Track/Experiment_Tracker.xlsx')
            modal_results = Modal(key="results",title='Experiment Results')
            res_col1, res_col2, res_col3 = st.sidebar.columns([0.28, 0.44, 0.28])

            with res_col2:
                open_results = st.button('Display Results :receipt:')
            if open_results:
                with modal_results.container():
                    st.write(ml_results)


                
        except:
            pass                
                





        







    

else:
    file_upload_options = ['Upload File', 'Use Default File']
    upload_options = st.selectbox('Select File Upload Option', file_upload_options)
    if upload_options == 'Upload File':
        uploaded_files = st.file_uploader("Upload a folder", type="csv", accept_multiple_files=False)
    else:
        uploaded_files = './upload_files_store/heart_2022_with_nans.csv'

    if uploaded_files is not None:
        df = pd.read_csv(uploaded_files)
        #st.write(len(df))
        df_dropped = df.dropna()
        #st.write(len(df_dropped))

        ############################ Select Label Column #########################
        column_names = df_dropped.columns
        
        if 'df_dropped_label' not in st.session_state:
            st.session_state.df_dropped_label = None

        if 'label_column_df' not in st.session_state:
            st.session_state.label_column_df = None

        if 'one_hot_encoded_data' not in st.session_state:
            st.session_state.one_hot_encoded_data = None

        if 'X_pca' not in st.session_state:
            st.session_state.X_pca = None

        if 'pca' not in st.session_state:
            st.session_state.pca = None

        if 'selected_task' not in st.session_state:
            st.session_state.selected_task = None

        if 'clf' not in st.session_state:
            st.session_state.clf = None

        if 'display_metrics_csv' not in st.session_state:
            st.session_state.display_metrics_csv = None

        drop_col1, drop_col2 = st.columns([0.75,0.25])
        with drop_col1:
            label_column = st.selectbox('**Select Label Column**', column_names)
            #
            # st.write(label_column_df)
        with drop_col2:
            st.write(' ')
            st.write(' ')
            if st.button('Select Label'):
                st.session_state.label_column_df = df_dropped.loc[:, str(label_column)]
                st.session_state.df_dropped_label = df_dropped.drop(str(label_column), axis = 'columns')
                
        
        
        try:
            if len(st.session_state.df_dropped_label) > 0:
                #st.write(st.session_state.label_column_df.shape)
                ############################# Select Task ##################################
                task = ['Classification', 'Regression']
                st.session_state.selected_task = st.selectbox('**Select Task to Perform**', task)
                if st.session_state.selected_task == 'Classification':
                    le = LabelEncoder()
                    Y = le.fit_transform(st.session_state.label_column_df)
                else:
                    Y = st.session_state.label_column_df

                ############################## Encoding ###########################
                categorical_columns = st.session_state.df_dropped_label.select_dtypes(include=['object']).columns.tolist()
                #st.write(len(categorical_columns))
                st.session_state.one_hot_encoded_data = pd.get_dummies(st.session_state.df_dropped_label, columns = categorical_columns)
                st.write(st.session_state.one_hot_encoded_data.head())

                ############################# Data Standardization ###############################
                scaler = StandardScaler()
                X_scaled = scaler.fit_transform(st.session_state.one_hot_encoded_data)
                ############################ Dimensionality Reduction #############################
                dr = ['Yes', 'No']
                dr_selected = st.selectbox('Perform Dimensionality Reduction', dr)
                if dr_selected == 'Yes':
                    ############################# PCA ###############################
                    pca_col1, pca_col2 = st.columns([0.7, 0.3])
                    with pca_col1:
                        no_pc = st.number_input('Select Number of Principal Components', min_value = 2, step = 1)
                    with pca_col2:
                        if st.button('Perform PCA'):
                            st.session_state.pca = PCA(n_components=no_pc)
                            st.session_state.pca.fit(X_scaled)
                            st.session_state.X_pca = st.session_state.pca.transform(X_scaled)   
                    try:
                        pass
                        #st.write("Explained Variance Ratio:", st.session_state.pca.explained_variance_ratio_)
                    except:
                        pass

                else:
                    st.session_state.X_pca = X_scaled
                    no_pc = 0

        except:
            pass

        if len(os.listdir('./Experiment_Track_CSV')) == 0:
            workbook = Workbook()
            worksheet = workbook.active

            col_names_list = ['No', 'Model Name', 'Accuracy', 
                                'Precision', 'Recall', 'f1 Score',
                                'Mean Absolute Error','Mean Squared Error', 'R2 Score']
            
            row_number = 1

            for col, value in enumerate(col_names_list, start=1):
                worksheet.cell(row=row_number, column=col, value=value)

            workbook.save(filename='./Experiment_Track_CSV/Experiment_Tracker.xlsx')

        if len(os.listdir('./Parameters_Track_CSV')) == 0:
                workbook = Workbook()
                worksheet = workbook.active

                col_names_list = ['No', 'Experiment Name', 'Label Column', 'Task', 
                                'Dimensionality Reduction', 'Principal Components', 'Test Split', 'Model']
                
                row_number = 1

                for col, value in enumerate(col_names_list, start=1):
                    worksheet.cell(row=row_number, column=col, value=value)

                workbook.save(filename='./Parameters_Track_CSV/Parameters_Tracker_CSV.xlsx')
        
        st.header('Train Test Split')
        test_split = st.number_input('test split', min_value = 0.1, max_value = 0.9, step = 0.01)
        if st.session_state.selected_task == 'Classification':
                classification_models = ['Logistic Regression', 'Decision Trees', 'Random Forest', 
                                        'XGBoost', 'Naive Bayes']
                classification_metrics = ['Accuracy', 'Precision', 'Recall', 'F1 Score']
                selected_model = st.selectbox('Select Classification Model', classification_models)
                selected_metrics = st.multiselect('Select Metrics', classification_metrics)
        elif st.session_state.selected_task == 'Regression':
                regression_models = ['Linear Regression', 'Polynomial Regression', 'Lasso Regression']
                selected_model = st.selectbox('Select Regression Model', regression_models)
                regression_metrics = ['Mean Absolute Error', 'Mean Squared Error', 'R2 Score']
                selected_metrics = st.multiselect('Select Metrics', regression_metrics)

        st.header('Experiment Name')
        if 'model_name' not in st.session_state:
            st.session_state.model_name = None
        st.session_state.model_name = st.text_input('Enter the experiment name')

        try:
            if st.button('Run Experiment'):
                if st.session_state.model_name:
                    workbook = load_workbook('./Experiment_Track_CSV/Experiment_Tracker.xlsx')
                    worksheet = workbook.active

                    non_empty_rows = 0

                    for row in worksheet.iter_rows():
                        if any(cell.value is not None for cell in row):
                            non_empty_rows += 1

                    col_names_list = []
                    X_train, X_test, y_train, y_test = train_test_split(st.session_state.X_pca, Y, test_size=float(test_split))
                    ################################# Classification ########################################
                    if selected_model == 'Logistic Regression':
                        st.session_state.clf = LogisticRegression()
                        st.session_state.clf.fit(X_train, y_train)
                    elif selected_model == 'Decision Trees':
                        st.session_state.clf = DecisionTreeClassifier()
                        st.session_state.clf.fit(X_train, y_train)
                    elif selected_model == 'Random Forest':
                        st.session_state.clf = RandomForestClassifier()
                        st.session_state.clf.fit(X_train, y_train)
                    elif selected_model == 'XGBoost':
                        st.session_state.clf = XGBClassifier()
                        st.session_state.clf.fit(X_train, y_train)
                    elif selected_model == 'Naive Bayes':
                        st.session_state.clf = GaussianNB()
                        st.session_state.clf.fit(X_train, y_train)
                    ################################# Regression ########################################
                    elif selected_model == 'Linear Regression':
                        st.session_state.clf = LinearRegression()
                        st.session_state.clf.fit(X_train, y_train)
                    elif selected_model == 'Polynomial Regression':
                        degree = st.number_input('Select Degree', min_value = 2, step = 1)
                        st.session_state.clf = PolynomialFeatures(degree = degree)
                        st.session_state.clf.fit_transform(X_train, y_train)
                    elif selected_model == 'Lasso Regression':
                        st.session_state.clf = Lasso()
                        st.session_state.clf.fit(X_train, y_train)
                    

                    filename = "./models_csv/" + str(st.session_state.model_name) + '.pkl'
                    # with open(filename, 'wb') as model_file:
                    joblib.dump(st.session_state.clf, filename)
                    #st.session_state.clf.save(filename)

                    y_pred = st.session_state.clf.predict(X_test)

                    col_names_list.append(non_empty_rows)
                    col_names_list.append(str(st.session_state.model_name))
                    
                    if 'Accuracy' in selected_metrics:
                        accuracy = accuracy_score(y_test, y_pred)
                        col_names_list.append(accuracy)
                    else:
                        col_names_list.append('Nan') 
                    if 'Precision' in selected_metrics:
                        precision = precision_score(y_test, y_pred)
                        col_names_list.append(precision)
                    else:
                        col_names_list.append('Nan')
                    if 'Recall' in selected_metrics:
                        recall = recall_score(y_test, y_pred)
                        col_names_list.append(recall)
                    else:
                        col_names_list.append('Nan')
                    if 'F1 Score' in selected_metrics:
                        f1 = f1_score(y_test, y_pred)
                        col_names_list.append(f1)
                    else:
                        col_names_list.append('Nan')
                    if 'Mean Absolute Error' in selected_metrics:
                        mae = mean_absolute_error(y_test, y_pred)
                        col_names_list.append(mae)
                    else:
                        col_names_list.append('Nan')
                    if 'Mean Squared Error' in selected_metrics:
                        mse = mean_squared_error(y_test, y_pred)
                        col_names_list.append(mse)
                    else:
                        col_names_list.append('Nan')
                    if 'R2 Score' in selected_metrics:
                        r2 = r2_score(y_test, y_pred)
                        col_names_list.append(r2)
                    else:
                        col_names_list.append('Nan')
                    # st.write('Accuracy', accuracy)
                    # st.write('Precision', precision)
                    # st.write('Recall', recall)
                    # st.write('F1 Score', f1)
                    for col, value in enumerate(col_names_list, start=1):
                        worksheet.cell(row=non_empty_rows+1, column=col, value=value)

                    workbook.save(filename='./Experiment_Track_CSV/Experiment_Tracker.xlsx')

                    ############################# Saving Parameters #########################################
                    workbook = load_workbook('./Parameters_Track_CSV/Parameters_Tracker_CSV.xlsx')
                    worksheet = workbook.active

                    non_empty_rows = 0

                    for row in worksheet.iter_rows():
                        if any(cell.value is not None for cell in row):
                            non_empty_rows += 1

                    col_params_list = []


                    col_params_list.append(non_empty_rows)
                    col_params_list.append(str(st.session_state.model_name))
                    col_params_list.append(label_column)
                    col_params_list.append(st.session_state.selected_task)
                    col_params_list.append(dr_selected)
                    col_params_list.append(no_pc)
                    col_params_list.append(test_split)
                    col_params_list.append(selected_model)


                    for col, value in enumerate(col_params_list, start=1):
                        worksheet.cell(row=non_empty_rows+1, column=col, value=value)
                    
                    workbook.save(filename='./Parameters_Track_CSV/Parameters_Tracker_CSV.xlsx')

                else:
                    st.write('Please provide name of the experiemnt to run the experiment')

            ##########################################################################################################################
                            

        except Exception as e:
            st.write(e)

        
        ################################### Sidebar ####################################################
        try:
            st.markdown(
                """
                <style>
                    section[data-testid="stSidebar"] {
                        width: 30% !important; # Set the width to your desired value
                    }
                </style>
                """,
                unsafe_allow_html=True,
            )
            with st.sidebar:
                st.header('Experiments')
                if len(os.listdir('./Parameters_Track_CSV')) == 0:
                    st.write('No experiments performed yet!!!')
                else:
                    wb_obj = load_workbook('./Parameters_Track_CSV/Parameters_Tracker_CSV.xlsx')
                    sheet_obj = wb_obj.active

                    total_rows = sheet_obj.max_row

                    
                    pop_up_names = []
                    pop_up_label = []
                    pop_up_task = []
                    pop_up_dr = []
                    pop_up_pc = []
                    pop_up_val_splits = []
                    pop_up_model = []

                    for u in range(2, total_rows+1):
                        pop_up_names.append(sheet_obj.cell(row = u, column = 2).value)
                        pop_up_label.append(sheet_obj.cell(row = u, column = 3).value)
                        pop_up_task.append(sheet_obj.cell(row = u, column = 4).value)
                        pop_up_dr.append(sheet_obj.cell(row = u, column = 5).value)
                        pop_up_pc.append(sheet_obj.cell(row = u, column = 6).value)
                        pop_up_val_splits.append(sheet_obj.cell(row = u, column = 7).value)
                        pop_up_model.append(sheet_obj.cell(row = u, column = 8).value)

                
                
                    pop_col1, pop_col2, pop_col3, pop_col4 = st.sidebar.columns([0.1, 0.4, 0.25, 0.25])
                    with pop_col1:
                        st.write('No')
                    with pop_col2:
                        st.write('Model Name')
                    with pop_col3:
                        st.write('Check')
                    with pop_col4:
                        st.write('Download')

                    
                    for y in range(len(pop_up_names)):
                        modal = Modal(key=str(y),title=str(pop_up_names[y]))
                        for i in range(1):
                            with pop_col1:
                                st.write(str(y+1))
                                st.write(' ')
                            with pop_col2:
                                st.write(pop_up_names[y])
                                st.write(' ')
                            with pop_col3:
                                open_modal = st.button('Show', key=str(y))
                            with pop_col4:
                                y_str = y+1000
                                don_filename = './models_csv/' + str(st.session_state.model_name) + '.pkl'
                                display_filename = str(st.session_state.model_name) + '.pkl'
                                with open(don_filename, 'rb') as f:
                                    st.download_button(':inbox_tray:', f, file_name=display_filename, key = str(y_str))  # :inbox_tray:  	:receipt:
                        
                        if open_modal:
                            with modal.container():
                                st.write('**Name**: ', pop_up_names[y])
                                st.write('**Label**: ', pop_up_label[y])
                                st.write('**Task: **', pop_up_task[y])
                                st.write('**Dimensionality Reduction: **', pop_up_dr[y])
                                st.write('**Number of Principal Components: **', pop_up_pc[y])
                                st.write('**Test Split: **', pop_up_val_splits[y])
                                st.write('**Model: **', pop_up_model[y])

            ml_results = pd.read_excel('./Experiment_Track_CSV/Experiment_Tracker.xlsx')
            modal_results = Modal(key="results",title='Experiment Results')
            res_col1, res_col2, res_col3 = st.sidebar.columns([0.28, 0.44, 0.28])

            with res_col2:
                open_results = st.button('Display Results :receipt:')
            if open_results:
                with modal_results.container():
                    st.write(ml_results)


                
        except:
            pass








